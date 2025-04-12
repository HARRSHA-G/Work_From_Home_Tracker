from django.db import models
from openpyxl import Workbook, load_workbook
import os

# Path to the Excel file
EXCEL_FILE = "work_from_home_data.xlsx"

class WorkFromHomeRecord(models.Model):
    user_id = models.CharField(max_length=100)
    work_location = models.CharField(max_length=100)
    address = models.TextField()
    ip_address = models.CharField(max_length=100)
    server_time = models.DateTimeField(auto_now_add=True)
    status = models.CharField(max_length=50, default="Present")

    def __str__(self):
        return f"{self.user_id} - {self.work_location}"

    def save(self, *args, **kwargs):
        # Save to the database
        super().save(*args, **kwargs)

        # Save to Excel file
        if not os.path.exists(EXCEL_FILE):
            wb = Workbook()
            ws = wb.active
            ws.append(["User ID", "Work Location", "Address", "IP Address", "Server Time", "Status"])
        else:
            wb = load_workbook(EXCEL_FILE)
            ws = wb.active

        ws.append([
            self.user_id,
            self.work_location,
            self.address,
            self.ip_address,
            self.server_time.strftime("%Y-%m-%d %H:%M:%S"),
            self.status,
        ])
        wb.save(EXCEL_FILE)
